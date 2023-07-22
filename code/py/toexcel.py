from cgi import test
import openpyxl
from datetime import datetime
import sys

fn = 'D:/專題/程式/test/data/DART.xlsx'
wb = openpyxl.load_workbook(fn)

wb.active = 0
ws = wb.active

localtime = datetime.now()
date = localtime.strftime("%Y-%m-%d")
time = localtime.strftime("%H:%M:%S")
#print(date, time)
if len(sys.argv) == 1:
    print('no argument', flush=True)
    sys.exit()
pass
doc_test = sys.argv[1]
temp = doc_test.split('@')
maxrow = ws.max_row+1

ws.cell(column=1, row=maxrow).value = date
ws.cell(column=2, row=maxrow).value = time
ws.cell(column=3, row=maxrow).value = temp[0]
ws.cell(column=4, row=maxrow).value = temp[1]
ws.cell(column=5, row=maxrow).value = temp[2]
ws.cell(column=6, row=maxrow).value = temp[3]
wb.save(fn)
